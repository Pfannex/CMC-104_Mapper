from openpyxl import load_workbook
import collections

def xls():
    wb = load_workbook(filename="CMC-104_Mapper/xls/test.xlsx")
    ws = wb.active
    lst = collections.defaultdict(list)
    for i in range(2,5):
        lst["Name"].append(ws.cell(i,1).value)
        lst["Alter"].append(ws.cell(i,2).value)
        
    return lst
